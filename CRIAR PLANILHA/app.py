# -- gerador de planilha de gestão financeira completa (pessoal)
from datetime import date
from pathlib import Path

#  pip install openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUT = Path(__file__).resolve().parent / "Gestao_Financeira_Pessoal.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Dashboard"

green = "2E7D32"
red = "C62828"
blue = "1565C0"
amber = "F9A825"
gray = "ECEFF1"
dark = "263238"
white = "FFFFFF"
light_green = "E8F5E9"
light_red = "FFEBEE"
light_blue = "E3F2FD"
light_amber = "FFF8E1"
border = Border(bottom=Side(style="thin", color="CFD8DC"))


def style_title(cell, size=18):
    cell.font = Font(bold=True, size=size, color=dark)
    cell.alignment = Alignment(horizontal="left")


def style_header(row):
    for c in row:
        c.font = Font(bold=True, color=white)
        c.fill = PatternFill("solid", fgColor=dark)
        c.alignment = Alignment(horizontal="center")


def style_money(cell, fill):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, size=12, color=dark)
    cell.alignment = Alignment(horizontal="center")
    cell.border = border
    cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'


def add_table(sheet, name, ref):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(tab)


def set_widths(sheet, widths):
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


months = [
    "Janeiro", "Fevereiro", "Marco", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]
expense_cats = [
    "Moradia", "Mercado", "Transporte", "Saude", "Educacao", "Lazer",
    "Assinaturas", "Contas", "Dividas", "Impostos", "Outros",
]
income_cats = ["Salario", "Freelance", "Bonus", "Investimentos", "Reembolso", "Outros"]
reserve_types = ["Reserva de emergencia", "Investimento", "Viagem", "Compra planejada", "Outro"]
payment_methods = ["Pix", "Debito", "Credito", "Dinheiro", "Boleto", "Transferencia"]
status_values = ["Pago", "Pendente", "Recebido", "Previsto"]

cat = wb.create_sheet("Categorias")
cat.sheet_state = "hidden"
lists = {
    "A": ("Meses", months),
    "B": ("CategoriasSaida", expense_cats),
    "C": ("CategoriasEntrada", income_cats),
    "D": ("TiposReserva", reserve_types),
    "E": ("FormasPagamento", payment_methods),
    "F": ("Status", status_values),
    "G": ("TipoLancamento", ["Entrada", "Saida"]),
    "H": ("Recorrencia", ["Fixo", "Variavel"]),
}
for col, (title, values) in lists.items():
    cat[f"{col}1"] = title
    for i, value in enumerate(values, 2):
        cat[f"{col}{i}"] = value

for title in [
    "Lancamentos", "Gastos Fixos", "Gastos Variaveis", "Ganhos Fixos",
    "Ganhos Variaveis", "Dinheiro Guardado",
]:
    wb.create_sheet(title)

lanc = wb["Lancamentos"]
lanc_headers = ["Data", "Mes", "Ano", "Tipo", "Recorrencia", "Categoria", "Descricao", "Forma", "Status", "Valor", "Observacao"]
lanc.append(lanc_headers)
sample = [
    [date(2026, 1, 5), "Janeiro", 2026, "Entrada", "Fixo", "Salario", "Salario mensal", "Transferencia", "Recebido", 0, ""],
    [date(2026, 1, 10), "Janeiro", 2026, "Saida", "Fixo", "Moradia", "Aluguel/financiamento", "Pix", "Pago", 0, ""],
    [date(2026, 1, 15), "Janeiro", 2026, "Saida", "Variavel", "Mercado", "Compras do mes", "Debito", "Pago", 0, ""],
]
for row in sample:
    lanc.append(row)
for r in range(5, 505):
    lanc[f"B{r}"] = f'=IF(A{r}="","",CHOOSE(MONTH(A{r}),"Janeiro","Fevereiro","Marco","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"))'
    lanc[f"C{r}"] = f'=IF(A{r}="","",YEAR(A{r}))'
for row in lanc.iter_rows(min_row=2, max_row=504, min_col=1, max_col=11):
    row[0].number_format = "dd/mm/yyyy"
    row[9].number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
add_table(lanc, "tblLancamentos", "A1:K504")
style_header(lanc[1])
set_widths(lanc, {"A": 13, "B": 12, "C": 10, "D": 12, "E": 13, "F": 18, "G": 28, "H": 16, "I": 12, "J": 14, "K": 28})
lanc.freeze_panes = "A2"

validations = [
    ("D2:D504", "Categorias!$G$2:$G$3"),
    ("E2:E504", "Categorias!$H$2:$H$3"),
    ("H2:H504", "Categorias!$E$2:$E$7"),
    ("I2:I504", "Categorias!$F$2:$F$5"),
]
for rng, formula in validations:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    lanc.add_data_validation(dv)
    dv.add(rng)

for sheet_name, table_name, headers, formulas in [
    ("Gastos Fixos", "tblGastosFixos", ["Categoria", "Descricao", "Vencimento", "Valor Previsto", "Valor Pago", "Status", "Observacao"], None),
    ("Gastos Variaveis", "tblGastosVariaveis", ["Categoria", "Descricao", "Mes", "Ano", "Valor Previsto", "Valor Real", "Diferenca", "Observacao"], "var_expense"),
    ("Ganhos Fixos", "tblGanhosFixos", ["Categoria", "Descricao", "Dia Recebimento", "Valor Previsto", "Valor Recebido", "Status", "Observacao"], None),
    ("Ganhos Variaveis", "tblGanhosVariaveis", ["Categoria", "Descricao", "Mes", "Ano", "Valor Previsto", "Valor Real", "Diferenca", "Observacao"], "var_income"),
]:
    s = wb[sheet_name]
    s.append(headers)
    for _ in range(2, 102):
        s.append([""] * len(headers))
    if formulas:
        for r in range(2, 102):
            s[f"G{r}"] = f'=IF(OR(E{r}="",F{r}=""),"",F{r}-E{r})'
    add_table(s, table_name, f"A1:{chr(64 + len(headers))}101")
    style_header(s[1])
    s.freeze_panes = "A2"
    set_widths(s, {"A": 18, "B": 28, "C": 16, "D": 12, "E": 16, "F": 16, "G": 16, "H": 28})
    for row in s.iter_rows(min_row=2, max_row=101):
        for cell in row:
            header = s.cell(1, cell.column).value
            if "Valor" in str(header) or header == "Diferenca":
                cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'

guard = wb["Dinheiro Guardado"]
guard_headers = ["Data", "Mes", "Ano", "Objetivo", "Tipo", "Valor", "Meta", "Progresso", "Onde esta guardado", "Observacao"]
guard.append(guard_headers)
for r in range(2, 202):
    guard[f"B{r}"] = f'=IF(A{r}="","",CHOOSE(MONTH(A{r}),"Janeiro","Fevereiro","Marco","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"))'
    guard[f"C{r}"] = f'=IF(A{r}="","",YEAR(A{r}))'
    guard[f"H{r}"] = f'=IF(G{r}=0,"",F{r}/G{r})'
for row in guard.iter_rows(min_row=2, max_row=201):
    row[0].number_format = "dd/mm/yyyy"
    row[5].number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
    row[6].number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
    row[7].number_format = "0%"
add_table(guard, "tblGuardado", "A1:J201")
style_header(guard[1])
set_widths(guard, {"A": 13, "B": 12, "C": 10, "D": 24, "E": 22, "F": 14, "G": 14, "H": 12, "I": 24, "J": 28})
guard.freeze_panes = "A2"
dv = DataValidation(type="list", formula1="Categorias!$D$2:$D$6", allow_blank=True)
guard.add_data_validation(dv)
dv.add("E2:E201")

for sheet_name in ["Gastos Fixos", "Gastos Variaveis"]:
    s = wb[sheet_name]
    dv = DataValidation(type="list", formula1="Categorias!$B$2:$B$12", allow_blank=True)
    s.add_data_validation(dv)
    dv.add("A2:A101")
for sheet_name in ["Ganhos Fixos", "Ganhos Variaveis"]:
    s = wb[sheet_name]
    dv = DataValidation(type="list", formula1="Categorias!$C$2:$C$7", allow_blank=True)
    s.add_data_validation(dv)
    dv.add("A2:A101")

dash = wb["Dashboard"]
dash["A1"] = "Gestao Financeira Pessoal"
style_title(dash["A1"], 20)
dash["A3"] = "Ano"
dash["B3"] = 2026
dash["C3"] = "Mes"
dash["D3"] = "Janeiro"
for c in ["A3", "C3"]:
    dash[c].font = Font(bold=True, color=dark)
dash["B3"].fill = PatternFill("solid", fgColor=gray)
dash["D3"].fill = PatternFill("solid", fgColor=gray)
dv_month = DataValidation(type="list", formula1="Categorias!$A$2:$A$13")
dash.add_data_validation(dv_month)
dv_month.add("D3")

cards = [
    ("A5", "Entradas do mes", "B5", '=SUMIFS(tblLancamentos[Valor],tblLancamentos[Tipo],"Entrada",tblLancamentos[Mes],$D$3,tblLancamentos[Ano],$B$3)', light_green),
    ("D5", "Saidas do mes", "E5", '=SUMIFS(tblLancamentos[Valor],tblLancamentos[Tipo],"Saida",tblLancamentos[Mes],$D$3,tblLancamentos[Ano],$B$3)', light_red),
    ("G5", "Guardado no mes", "H5", '=SUMIFS(tblGuardado[Valor],tblGuardado[Mes],$D$3,tblGuardado[Ano],$B$3)', light_blue),
    ("A8", "Saldo livre do mes", "B8", "=B5-E5-H5", light_amber),
    ("D8", "Total guardado", "E8", "=SUM(tblGuardado[Valor])", light_blue),
    ("G8", "Meta total", "H8", "=SUM(tblGuardado[Meta])", light_green),
]
for label_cell, label, value_cell, formula, fill in cards:
    dash[label_cell] = label
    dash[label_cell].font = Font(bold=True, color=dark)
    dash[value_cell] = formula
    style_money(dash[value_cell], fill)

dash["A11"] = "Progresso das reservas"
dash["B11"] = '=IF(H8=0,0,E8/H8)'
dash["B11"].number_format = "0%"
dash["B11"].fill = PatternFill("solid", fgColor=light_blue)
dash["B11"].font = Font(bold=True, color=dark)

dash["A14"] = "Resumo mensal"
style_title(dash["A14"], 14)
summary_headers = ["Mes", "Entradas", "Saidas", "Guardado", "Saldo Livre"]
for col, header in enumerate(summary_headers, 1):
    dash.cell(15, col).value = header
style_header(dash[15][:5])
for i, month in enumerate(months, 16):
    dash.cell(i, 1).value = month
    dash.cell(i, 2).value = f'=SUMIFS(tblLancamentos[Valor],tblLancamentos[Tipo],"Entrada",tblLancamentos[Mes],A{i},tblLancamentos[Ano],$B$3)'
    dash.cell(i, 3).value = f'=SUMIFS(tblLancamentos[Valor],tblLancamentos[Tipo],"Saida",tblLancamentos[Mes],A{i},tblLancamentos[Ano],$B$3)'
    dash.cell(i, 4).value = f'=SUMIFS(tblGuardado[Valor],tblGuardado[Mes],A{i},tblGuardado[Ano],$B$3)'
    dash.cell(i, 5).value = f"=B{i}-C{i}-D{i}"
    for col in range(2, 6):
        dash.cell(i, col).number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'

dash["G14"] = "Saidas por categoria"
style_title(dash["G14"], 14)
dash["G15"] = "Categoria"
dash["H15"] = "Valor"
style_header([dash["G15"], dash["H15"]])
for i, cat_name in enumerate(expense_cats, 16):
    dash.cell(i, 7).value = cat_name
    dash.cell(i, 8).value = f'=SUMIFS(tblLancamentos[Valor],tblLancamentos[Tipo],"Saida",tblLancamentos[Categoria],G{i},tblLancamentos[Mes],$D$3,tblLancamentos[Ano],$B$3)'
    dash.cell(i, 8).number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'

bar = BarChart()
bar.title = "Entradas x Saidas x Guardado"
bar.y_axis.title = "Valor"
bar.x_axis.title = "Mes"
bar.add_data(Reference(dash, min_col=2, max_col=4, min_row=15, max_row=27), titles_from_data=True)
bar.set_categories(Reference(dash, min_col=1, min_row=16, max_row=27))
bar.height = 7
bar.width = 15
dash.add_chart(bar, "A30")

dough = DoughnutChart()
dough.title = "Saidas por categoria"
dough.add_data(Reference(dash, min_col=8, min_row=15, max_row=26), titles_from_data=True)
dough.set_categories(Reference(dash, min_col=7, min_row=16, max_row=26))
dough.height = 7
dough.width = 11
dash.add_chart(dough, "G30")

for row in dash.iter_rows(min_row=1, max_row=45, min_col=1, max_col=10):
    for cell in row:
        cell.alignment = Alignment(vertical="center")
for col in range(1, 11):
    dash.column_dimensions[chr(64 + col)].width = 17
dash.column_dimensions["A"].width = 20
dash.column_dimensions["G"].width = 22
dash.freeze_panes = "A14"

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

for s in [lanc, guard, wb["Gastos Fixos"], wb["Gastos Variaveis"], wb["Ganhos Fixos"], wb["Ganhos Variaveis"]]:
    target = "J2:J504" if s.title == "Lancamentos" else "F2:F201"
    s.conditional_formatting.add(
        target,
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color=red)),
    )

wb.save(OUT)
print(f"Planilha criada em: {OUT}")
